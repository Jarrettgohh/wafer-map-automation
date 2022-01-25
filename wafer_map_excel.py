import json
import os

from openpyxl.utils.cell import column_index_from_string
from openpyxl.chart.shapes import GraphicalProperties
from openpyxl.chart.marker import Marker, DataPoint
from openpyxl.drawing.fill import PatternFillProperties, ColorChoice
from openpyxl import load_workbook
from openpyxl.chart import (
    ScatterChart,
    Reference,
    Series,
)

f = open('config.json')
config_json = json.load(f)

wafer_mapping_configurations = config_json['wafer_mapping_configurations']
to_plot = wafer_mapping_configurations['to_plot']
area_fraction = wafer_mapping_configurations['area_fraction']
color_indicators = wafer_mapping_configurations['color_indicators']

to_plot_rows = to_plot['rows']
to_plot_columns = to_plot['columns']

filename = 'wafer-mapping-automation-test.xlsx'
file_dir = f'C:/Users/gohja/Desktop/wafer-map-automation/{filename}'

#
# Open the excel workbook
#

wb = load_workbook(filename=filename)
ws = wb.active

# Create a scatter chart
chart = ScatterChart()
chart.title = "Scatter Chart Automation Test"
chart.legend = None

x_col = column_index_from_string(to_plot_columns[0])
y_col = column_index_from_string(to_plot_columns[1])
min_row = to_plot_rows[0]
max_row = to_plot_rows[1]

xvalues = Reference(ws, min_col=x_col, min_row=min_row, max_row=max_row)
yvalues = Reference(ws, min_col=y_col, min_row=min_row, max_row=max_row)

# Plot the points on the scatter chart
series = Series(xvalues, yvalues)
# chart.add_data(xvalues, yvalues)

# series.marker = DataPoint(
#     idx=5, spPr=GraphicalProperties(solidFill="800000"))

series.marker = Marker('circle', size=15)
series.graphicalProperties.line.noFill = True
# s.marker.graphicalProperties.solidFill = "800000"

pt = DataPoint(idx=10)
pt.marker = Marker('circle', size=15)
pt.graphicalProperties.line.noFill = True

# pt = DataPoint(idx=10,
#                marker=Marker('circle', size=15),
#                spPr=GraphicalProperties(solidFill="800000"))

pt = DataPoint(idx=5)
pt.graphicalProperties.solidFill = "800000"

series.dPt.append(pt)

chart.series.append(series)

ws.add_chart(chart, "G14")

#
# Calculation of the area fraction, and coloring the markers on the chart
#

# area_fraction_rows = area_fraction['rows']
# area_fraction_col = area_fraction['column']

# for cell_row_index in range(area_fraction_rows[0], area_fraction_rows[1] + 1):
#     area_fraction = ws[area_fraction_col + str(cell_row_index)].value
#     area_fraction_percentage = area_fraction * 100

#     for color_indicator in color_indicators:

wb.save(filename)
os.startfile(file_dir)