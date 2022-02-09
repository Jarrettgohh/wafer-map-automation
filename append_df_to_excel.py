from pathlib import Path
from copy import copy
from typing import Union, Optional
import numpy as np
import pandas as pd
import openpyxl

from openpyxl.styles import Alignment
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter


def copy_excel_cell_range(
        src_ws: openpyxl.worksheet.worksheet.Worksheet,
        min_row: int = None,
        max_row: int = None,
        min_col: int = None,
        max_col: int = None,
        tgt_ws: openpyxl.worksheet.worksheet.Worksheet = None,
        tgt_min_row: int = 1,
        tgt_min_col: int = 1,
        with_style: bool = True) -> openpyxl.worksheet.worksheet.Worksheet:

    if tgt_ws is None:
        tgt_ws = src_ws

    # https://stackoverflow.com/a/34838233/5741205
    for row in src_ws.iter_rows(min_row=min_row,
                                max_row=max_row,
                                min_col=min_col,
                                max_col=max_col):
        for cell in row:
            tgt_cell = tgt_ws.cell(row=cell.row + tgt_min_row - 1,
                                   column=cell.col_idx + tgt_min_col - 1,
                                   value=cell.value)

            if with_style and cell.has_style:
                tgt_cell._style = copy(cell._style)
                tgt_cell.font = copy(cell.font)
                tgt_cell.border = copy(cell.border)
                tgt_cell.fill = copy(cell.fill)
                tgt_cell.number_format = copy(cell.number_format)
                tgt_cell.protection = copy(cell.protection)
                tgt_cell.alignment = copy(cell.alignment)
    return tgt_ws


def append_df_to_excel(filename: Union[str, Path],
                       df: pd.DataFrame,
                       sheet_name: str = 'Sheet1',
                       startrow: Optional[int] = None,
                       startcol: Optional[int] = None,
                       max_col_width: int = 30,
                       autofilter: bool = False,
                       fmt_int: str = "#,##0",
                       fmt_float: str = "#,##0.00",
                       fmt_date: str = "yyyy-mm-dd",
                       fmt_datetime: str = "yyyy-mm-dd hh:mm",
                       truncate_sheet: bool = False,
                       storage_options: Optional[dict] = None,
                       **to_excel_kwargs) -> None:

    def set_column_format(ws, column_letter, fmt):
        for cell in ws[column_letter]:
            cell.number_format = fmt

    filename = Path(filename)
    file_exists = filename.is_file()

    # process parameters
    # calculate first column number
    # if the DF will be written using `index=True`, then `first_col = 2`, else `first_col = 1`
    # first_col = int(to_excel_kwargs.get("index", True)) + 1

    # ignore [engine] parameter if it was passed
    if 'engine' in to_excel_kwargs:
        to_excel_kwargs.pop('engine')
    # save content of existing sheets
    if file_exists:
        wb = load_workbook(filename)
        sheet_names = wb.sheetnames
        sheet_exists = sheet_name in sheet_names
        sheets = {ws.title: ws for ws in wb.worksheets}

    with pd.ExcelWriter(filename.with_suffix(".xlsx"),
                        engine="openpyxl",
                        mode="a" if file_exists else "w",
                        if_sheet_exists="new" if file_exists else None,
                        date_format=fmt_date,
                        datetime_format=fmt_datetime,
                        storage_options=storage_options) as writer:
        if file_exists:
            # try to open an existing workbook
            writer.book = wb
            # get the last row in the existing Excel sheet
            # if it was not specified explicitly
            if startrow is None and sheet_name in writer.book.sheetnames:
                startrow = writer.book[sheet_name].max_row
            # truncate sheet
            if truncate_sheet and sheet_name in writer.book.sheetnames:
                # index of [sheet_name] sheet
                idx = writer.book.sheetnames.index(sheet_name)
                # remove [sheet_name]
                writer.book.remove(writer.book.worksheets[idx])
                # create an empty sheet [sheet_name] using old index
                writer.book.create_sheet(sheet_name, idx)
            # copy existing sheets
            writer.sheets = sheets
        else:
            # file doesn't exist, we are creating a new one
            startrow = 0

        # write out the DataFrame to an ExcelWriter
        df.to_excel(writer,
                    sheet_name=sheet_name,
                    startcol=startcol,
                    **to_excel_kwargs)
        worksheet = writer.sheets[sheet_name]

        if autofilter:
            worksheet.auto_filter.ref = worksheet.dimensions

        for x_col_no, dtyp in enumerate(df.dtypes, startcol + 1):

            column_letter = get_column_letter(x_col_no)

            if np.issubdtype(dtyp, np.integer):
                set_column_format(worksheet, column_letter, fmt_int)
            if np.issubdtype(dtyp, np.floating):
                set_column_format(worksheet, column_letter, fmt_float)

        # Wrap the text
        for row in worksheet.rows:
            for cell in row:
                cell.alignment = Alignment(wrap_text=True, vertical='center')

    if file_exists and sheet_exists:
        # move (append) rows from new worksheet to the `sheet_name` worksheet
        wb = load_workbook(filename)
        # retrieve generated worksheet name
        new_sheet_name = set(wb.sheetnames) - set(sheet_names)
        if new_sheet_name:
            new_sheet_name = list(new_sheet_name)[0]

        # copy rows written by `df.to_excel(...)` to
        copy_excel_cell_range(src_ws=wb[new_sheet_name],
                              tgt_ws=wb[sheet_name],
                              tgt_min_row=startrow + 1,
                              with_style=True)

        # remove new (generated by Pandas) worksheet
        del wb[new_sheet_name]
        wb.save(filename)
        wb.close()
