import json
import os
import sys
import pandas as pd

from openpyxl.utils.cell import column_index_from_string
from openpyxl.styles import PatternFill
from openpyxl.chart.shapes import GraphicalProperties
from openpyxl.chart.marker import Marker
from openpyxl import load_workbook
from openpyxl.chart import (
    ScatterChart,
    Reference,
    Series,
)

from append_df_to_excel import append_df_to_excel
from functions import pretty_print, pretty_print_error_msg

f = open('config.json')
config_json = json.load(f)

html_file_directory = config_json['html_file_directory']
wafer_mapping_configurations = config_json['wafer_mapping_configurations']

file_directory = wafer_mapping_configurations['file_directory']
file_name = wafer_mapping_configurations['file_name']

wafer_information = wafer_mapping_configurations['wafer_information']
number_of_wafer_points = wafer_information['number_of_wafer_points']
wafer_ids = wafer_information['wafer_ids']

error_information = wafer_mapping_configurations['error_information']
error_information_cell_color = error_information['cell_color']
error_information_scatter_site_color = error_information['scatter_site_color']

to_plot = wafer_mapping_configurations['to_plot']
to_plot_rows = to_plot['rows']
to_plot_columns = to_plot['columns']

area_fraction = wafer_mapping_configurations['area_fraction']
area_fraction_rows = area_fraction['rows']
area_fraction_columns = area_fraction['columns']

color_indicators = wafer_mapping_configurations['color_indicators']

#
# Open the excel workbook
#

file_path = f'{file_directory}/{file_name}'


def write_area_fraction_to_excel(site_defect_fraction_data: list):

    pretty_print(f'Compiling data in the excel file at path: {file_path}')

    len_site_defect_fraction_data = len(site_defect_fraction_data)
    len_wafer_ids = len(wafer_ids)

    if (number_of_wafer_points *
            len_wafer_ids) != len_site_defect_fraction_data:

        pretty_print_error_msg(
            f'`The product of the number_of_wafer_points` and the number of items in `wafer_ids` should equal to the number of images provided in {html_file_directory}]'
        )
        sys.exit()

    try:

        to_write_col = column_index_from_string(
            area_fraction_columns['area_fraction']) - 1

        for wafer_batch_index, wafer_id in enumerate(wafer_ids):
            pretty_print(f'Working on {wafer_id}')

            site_defect_fraction_data_start_index_to_read = wafer_batch_index * number_of_wafer_points

            # Iterate site defect fraction data
            for site_defect_fraction in site_defect_fraction_data[
                    site_defect_fraction_data_start_index_to_read:
                    site_defect_fraction_data_start_index_to_read +
                    number_of_wafer_points]:

                # Failed to read from the image
                if site_defect_fraction == None:
                    continue

                site_number = int(site_defect_fraction['site'])
                defect_fraction = float(
                    site_defect_fraction['defect_fraction'])

                print(
                    f'Appending defect fraction data for site {site_number}...'
                )

                row_to_write = area_fraction_rows[1] - (
                    number_of_wafer_points - site_number) - 1
                df = pd.DataFrame([defect_fraction])

                append_df_to_excel(
                    filename=file_path,
                    df=df,
                    sheet_name=wafer_id,
                    startrow=row_to_write,
                    startcol=to_write_col,
                )

            # Plot scatter graph
            plot_scatter_graph(sheet_name=wafer_id)

    except PermissionError:
        pretty_print_error_msg(
            f'Failed to write excel file at path: {file_path}. Do ensure that the file is not open/running.'
        )
        sys.exit()

    except Exception as e:
        pretty_print_error_msg(
            f'Something went wrong with writing to the excel file at path: {file_path}'
        )
        print(e)

        sys.exit()


def plot_scatter_graph(sheet_name: str):

    print(f'Plotting the scatter graph for {sheet_name}...')

    wb = load_workbook(filename=file_path)
    ws = wb[sheet_name]

    # Create a scatter chart
    chart = ScatterChart()
    chart.title = sheet_name
    chart.legend = None

    x_col = column_index_from_string(to_plot_columns['X_axis'])
    y_col = column_index_from_string(to_plot_columns['Y_axis'])
    min_row = to_plot_rows[0]
    max_row = to_plot_rows[1]

    for plot_row_index in range(min_row, max_row + 1):

        area_fraction_value = ws[area_fraction_columns['area_fraction'] +
                                 str(plot_row_index)].value

        is_area_fraction_value_undefined = area_fraction_value == None

        error_cell_color = None
        cell_color = None

        # There is no area/defect fraction value
        if is_area_fraction_value_undefined:
            error_cell_color = error_information_cell_color if error_information_cell_color != None else "800000"
            scatter_site_color = error_information_scatter_site_color

        else:
            scatter_site_color = '4472C4'
            area_fraction_percentage = area_fraction_value * 100

            ws[area_fraction_columns['area_fraction_percentage'] +
               str(plot_row_index)].value = area_fraction_percentage

            for color_indicator in color_indicators:
                color_indicator_range = color_indicators[color_indicator]
                color_indicator_lower_range = color_indicator_range[0]
                color_indicator_upper_range = color_indicator_range[1]

                is_area_fraction_percentage_in_range = area_fraction_percentage >= color_indicator_lower_range and area_fraction_percentage <= color_indicator_upper_range

                if is_area_fraction_percentage_in_range:
                    scatter_site_color = color_indicator
                    break

            cell_color = scatter_site_color

        # Color the X and Y axis cells and the area fraction percentage cells with color
        if cell_color != None:
            cellFill = PatternFill(fill_type='solid', fgColor=cell_color)

            ws[to_plot_columns['X_axis'] + str(plot_row_index)].fill = cellFill
            ws[to_plot_columns['Y_axis'] + str(plot_row_index)].fill = cellFill
            ws[area_fraction_columns['area_fraction_percentage'] +
               str(plot_row_index)].fill = cellFill

        # Color the cell under `area_fraction` with error color
        if error_cell_color != None:
            cellFill = PatternFill(fill_type='solid', fgColor=error_cell_color)

            ws[to_plot_columns['X_axis'] + str(plot_row_index)].fill = cellFill
            ws[to_plot_columns['Y_axis'] + str(plot_row_index)].fill = cellFill
            ws[area_fraction_columns['area_fraction'] +
               str(plot_row_index)].fill = cellFill
            ws[area_fraction_columns['area_fraction_percentage'] +
               str(plot_row_index)].fill = cellFill

        xvalues = Reference(ws, min_col=x_col, min_row=plot_row_index)
        yvalues = Reference(ws, min_col=y_col, min_row=plot_row_index)

        if scatter_site_color == None:
            continue

        # Plot and color the points on the scatter chart
        series = Series(xvalues, yvalues)
        series.marker = Marker(
            'circle',
            size=15,
            spPr=GraphicalProperties(solidFill=scatter_site_color))
        series.graphicalProperties.line.noFill = True

        chart.series.append(series)

    ws.add_chart(chart, "G14")

    try:
        wb.save(file_name)

    except PermissionError:
        pretty_print_error_msg(
            f'Failed to save excel file at path: {file_path}. Do ensure that the file is not open/running.'
        )
        sys.exit()


def wafer_map_excel(site_defect_fraction_data: list):
    try:
        write_area_fraction_to_excel(site_defect_fraction_data)
        os.startfile(file_path)

    except KeyboardInterrupt:
        sys.exit()
